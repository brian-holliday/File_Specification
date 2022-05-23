# import libraries
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import os
from datetime import date

# Functions

def get_review_file_path(site, repo_config_file):
    
    return repo_config_file.loc[repo_config_file["Site Name"] == sjc_sites_dict[site], 
                                "Review File Path"].values[0]
def set_file_path():
    
    file_path = str(input("Please point to file path for external request: "))
    return file_path

def site_file_path(): # Function to get data for external request

    
    file_path = set_file_path()

    no_sjc_sites = False

    try:
        sites = os.listdir(file_path)
        sites = [site for site in sites if site in sjc_sites]
        
        if len(sites) == 0:
            print("There are no SJC Sites in this directory\n")
            no_sjc_sites = True
            exit() # This exit statement will put you in the except clause
        else:
            return file_path
        
    except:
        if no_sjc_sites == True:
            exit()

        else:
            print("Could not find file path")
            exit()

        

def populate_data_df(path):

    data_list = []
    original_path = path

    sjc_sites = ["Ada", "Allegheny", "Buncombe", "Charleston", "Cook", "Harris", "Lucas",
                "Mecklenburg", "Milwaukee", "Multnomah", "New Orleans", "Palm Beach County",
                "Pennington", "Philadelphia", "Pima", "Spokane", "St. Louis", "San Francisco", 
                "East Baton Rouge", "Lake", "Minnehaha", "Missoula"]

    sites = os.listdir(original_path)
    sites = [site for site in sites if site in sjc_sites] # make sure folder of sjc sites only

    for site in sites:
        site_path = original_path + "\\" + site # look into the path of each site to get system points
        data_category = os.listdir(site_path)
        for category in data_category:
            if len(category.split(".")) > 1: # To make sure to get all folders and not files
                next
            else:
                current_category = category
                sub_category_path = site_path + "\\" + current_category
                sub_category = os.listdir(sub_category_path)
                for sub in sub_category:
                    if len(sub.split(".")) > 1:
                        next
                    else:
                        current_sub = sub
                        inclusion_years_path = sub_category_path + "\\" + current_sub
                        inclusion_years = os.listdir(inclusion_years_path)
                        for year in inclusion_years:
                            if len(year.split(".")) > 1:
                                next
                            else:
                                current_year = year
                                files_path = inclusion_years_path + "\\" + current_year 
                                files = os.listdir(files_path)
                                for file in files: 
                                    row_dict = {"Site":site, "Data_Category":category, 
                                                "Sub_Category":sub, "Inclusion Year":year, 
                                                "File":file}
                                    data_list.append(row_dict)
    return pd.DataFrame(data_list)

# Function to make sure that we have the right review file...
# For File Descriptions and Variable Listings
def get_review_file(review_file_list, review_file_summary, review_file_listings, site):
    if len(review_file_list) == 0:
        review_file_list.append(site)
    
    elif site != review_file_list[-1]:
        review_file_list.append(site)
        
        review_file_summary = pd.read_excel(get_review_file_path(site, repo_config), # pass in the config file
                                    sheet_name = "Summary", skiprows = 2, 
                                    names = ["System Point", "Sub-System Point", "Inclusion Year(s)", 
                                              "Pre-Processed Path", "Pre-Processed File Name", 
                                              "Processed Path", "Processed File Name",
                                              "Total Variables", "Total # of Variables Scrambled (only)"])

        review_file_listings = pd.read_excel(get_review_file_path(site, repo_config), # pass in the config file
                                    sheet_name = "Supplemental Listings", skiprows = 1, 
                                    names = ["Variable Name", "Variable Type", "System Point", 
                                              "Sub-System Point", "Inclusion Year(s)", 
                                              "File Path", "File Name", "Action Taken"])        
    else: # This is if the review file site does not need to change...will return same review file passed in.
        review_file_list.append(site)
        
    return review_file_list, review_file_summary, review_file_listings

# Function to populate file descriptions and variable listings dataframes
def get_descriptions_and_listings(data_df):
    # list to keep track of what reviw file we are using
    review_file_list = []
    
    # import initial review file summary sheet
    review_file_summary = pd.read_excel(get_review_file_path(data_df.loc[0, "Site"], repo_config), # pass in the config file
                                    sheet_name = "Summary", skiprows = 2, 
                                    names = ["System Point", "Sub-System Point", "Inclusion Year(s)", 
                                              "Pre-Processed Path", "Pre-Processed File Name", 
                                              "Processed Path", "Processed File Name",
                                              "Total Variables", "Total # of Variables Scrambled (only)"])
    # import inital review file listings sheet
    review_file_listings = pd.read_excel(get_review_file_path(data_df.loc[0, "Site"], repo_config), # pass in the config file
                                    sheet_name = "Supplemental Listings", skiprows = 1, 
                                    names = ["Variable Name", "Variable Type", "System Point", 
                                              "Sub-System Point", "Inclusion Year(s)", 
                                              "File Path", "File Name", "Action Taken"]) 
    
    file_descriptions = pd.read_excel("R:\SJC\Site information\Data Processing\Data Repository\Templates\External Data Request - File Specification Document - Template.xlsx",
                                     sheet_name = "File Descriptions")
    
    variable_listings = pd.read_excel("R:\SJC\Site information\Data Processing\Data Repository\Templates\External Data Request - File Specification Document - Template.xlsx",
                                     sheet_name = "Variable Listings", skiprows = 2,
                                     names = ["Site", "Data Category", "Data Subcategory", "Inclusion Year(s)",
                                              "File Name", "Name", "Type", "Indentifier Indicator (Y/N)"]) 
    
    for index, row in data_df.iterrows(): # Loop to go through files in external request
        # To make sure that we have the right review file
        review_file_list, review_file_summary, review_file_listings = get_review_file(review_file_list, review_file_summary, 
                                                                                      review_file_listings, row["Site"])
        
        review_file_summary_row_join = review_file_summary.loc[(review_file_summary["System Point"] == row["Data_Category"]) & 
                                                      (review_file_summary["Sub-System Point"] == row["Sub_Category"]) & 
                                                      (review_file_summary["Inclusion Year(s)"] == row["Inclusion Year"]), ].copy()
        
        # Call check file summary function to make sure that we are getting the correct rows 
        # review_file_summary_row_join = check_file_summary(data_df, review_file_summary_row_join, row["Site"])

        
        review_file_listings_row_join = review_file_listings.loc[(review_file_listings["System Point"] == row["Data_Category"]) &
                                                        (review_file_listings["Sub-System Point"] == row["Sub_Category"]) &
                                                        (review_file_listings["Inclusion Year(s)"] == row["Inclusion Year"]), ].copy()
        
        # review_file_listings_row_join = check_file_listings(data_df, review_file_listings_row_join, row["Site"])
        
        # Define Site Variable
        review_file_summary_row_join["Site"] = row["Site"]
        review_file_listings_row_join["Site"] = row["Site"]
        
        # Add information from the review file summary to the file_desciptions dataframe
        review_file_summary_row_join = review_file_summary_row_join[["Site", "System Point", "Sub-System Point", 
                                                   "Inclusion Year(s)", "Processed Path"]]
        
        review_file_summary_row_join.columns = file_descriptions.columns.tolist()
        
        file_descriptions = pd.concat([file_descriptions, review_file_summary_row_join], axis = 0)
        
        # Add information from review file listings to the variable listings dataframe 
        review_file_listings_row_join = review_file_listings_row_join[["Site", "System Point", "Sub-System Point",
                                                     "Inclusion Year(s)", "File Path",
                                                     "Variable Name", "Variable Type", 
                                                     "Action Taken"]]
        
        review_file_listings_row_join.columns = variable_listings.columns.tolist()
        
        variable_listings = pd.concat([variable_listings, review_file_listings_row_join], axis = 0)
        
    # Drop Duclicate rows
    file_descriptions = file_descriptions.drop_duplicates()
    variable_listings = variable_listings.drop_duplicates()
    
    # Edit the Identifier Indicator Variable in the variable listings sheet
    variable_listings = variable_listings.replace(to_replace = [np.nan, "Extract Year", 'Strip'],
                                                  value = "N")
    
    variable_listings = variable_listings.replace(to_replace = "Scrambled",
                                                  value = "Y")
    
    file_descriptions = get_proper_file_names(file_descriptions)
    variable_listings = get_proper_file_names(variable_listings)
    
    return file_descriptions, variable_listings

def get_proper_file_names(df):
    
    df = df.reset_index().drop("index", axis = 1)
    
    df["Actual File Names"] = np.nan
    
    for index, row in df.iterrows():
        df.loc[index, "Actual File Names"] = str(row["File Name"]).split("\\")[-1]
        
    df["File Name"] = df["Actual File Names"]
    
    df = df.drop("Actual File Names", axis = 1)
    
    return df 

def clean_descriptions_and_listings(df):
    
    df["File Name Check"] = np.nan
    
    for index, row in data_df.iterrows():
        row_mask = df["File Name"] == row["File"]
        
        if sum(row_mask) == 0:
            next
        else:
            df.loc[row_mask, "File Name Check"] = "Here"
    
    not_here_mask = df["File Name Check"] != "Here"
            
    df = df.loc[~not_here_mask, ]
    df = df.reset_index().drop(["index", "File Name Check"], axis = 1)
    
    return df

def check_template(file_path): # Function to check if template of the same name already exists
    return os.path.exists(file_path)

def get_final_save_path(file_path): #Function to get the final save destination for template
        
    file_path_list = file_path.split("\\")
    
    file_path_name = str(file_path_list[-2] + " - File Specification Document - " +  
                         date.today().strftime("%m-%d-%Y") + ".xlsx") 
    
    file_path_list.append(file_path_name)
        
    save_path = "\\"
    
    save_path = save_path.join(file_path_list)
    
    file_specs_empty_template = openpyxl.load_workbook("R:\\SJC\\Site information\\Data Processing\\Data Repository\\Templates\\External Data Request - File Specification Document - Template v2.xlsx")
    template_exists = check_template(save_path) # check to see if file with the same name exists
    
    if template_exists == True:
        os.remove(save_path)
    
    try:
        file_specs_empty_template.save(save_path) # save empty template in external requests folder.
        file_specs_empty_template.close()
    
    except:
        print("Could not save empty template in directory")
        exit()
        
    return save_path


def populate_specs_workbook(file_descriptions, variable_listings, file_path):
    
    file_path = get_final_save_path(file_path)
    
    file_specs = openpyxl.load_workbook(file_path)
    
    writer = pd.ExcelWriter(file_path, engine = "openpyxl")
    
    writer.book = file_specs # get openpyxl workbook object
    
    writer.sheets = {ws.title: ws for ws in file_specs.worksheets}
    
    for sheetname in writer.sheets:
        if sheetname == "File Descriptions":
            file_descriptions.to_excel(writer, sheet_name = sheetname, startrow = 1, startcol = 0,
                                      index = False, header = False)
        elif sheetname == "Variable Listings":
            variable_listings.to_excel(writer, sheet_name = sheetname, startrow = 2, startcol = 0,
                                      index = False, header = False)
        elif sheetname == "General Notes":
            sheet = file_specs[sheetname]
            sheet["A21"] = "File Last Updated: " + date.today().strftime("%m-%d-%Y")
        else:
            next # skip if it is not the right sheet.
    
    writer.save() # save the template
    
    return file_path

def format_workbook_columns(file_path):
    
    file_specs = openpyxl.load_workbook(file_path) # open the template
    
    writer = pd.ExcelWriter(file_path, engine = "openpyxl")
    
    writer.book = file_specs # get openpyxl workbook object
    
    writer.sheets = {ws.title: ws for ws in file_specs.worksheets}
    
    center_alignment = Alignment(horizontal = "center",
                                        vertical = "center")

            
    right_alignment = Alignment(horizontal = "right",
                                vertical = "center")
    # Format site column
    non_bold_font = Font(name = "Calibri",
                         size = 12,
                         bold = False,
                         italic = False,
                         vertAlign = None,
                         underline = "none",
                         strike = False)
            
    non_bold_italic_font = Font(name = "Calibri",
                                size = 12,
                                bold = False,
                                italic = True,
                                vertAlign = None,
                                underline = "none",
                                strike = False)
            
    bold_font = Font(name = "Calibri",
                     size = 12,
                     bold = True,
                     italic = False,
                     vertAlign = None,
                     underline = "none",
                     strike = False)
                
    for sheetname in writer.sheets:
        if sheetname == "File Descriptions":
            descriptions = file_specs[sheetname] # get the file_descriptions sheet
            
            # format columns
            
            # site
            site_column = descriptions["A1"]
            site_column.font = bold_font
            site_column.alignment = center_alignment
            
            # Data Category
            data_category_column = descriptions["B1"]
            data_category_column.font = bold_font
            data_category_column.alignment = center_alignment
            
            # Data Sub Category
            data_subcategory_column = descriptions["C1"]
            data_subcategory_column.font = non_bold_italic_font
            data_subcategory_column.alignment = center_alignment
            
            # Inclusion Year(s)
            inclusion_years_column = descriptions["D1"]
            inclusion_years_column.font = non_bold_italic_font
            inclusion_years_column.alignment = center_alignment
            
            # File Name
            file_name_column = descriptions["E1"]
            file_name_column.font = bold_font
            file_name_column.alignment = center_alignment
            
            for row in descriptions[2:descriptions.max_row]:  # skip the header
                site_cell = row[0]    # site column
                site_cell.alignment = center_alignment
                site_cell.font = bold_font
            
                data_category_cell = row[1] # data category column
                data_category_cell.font = bold_font # edit font
                data_category_cell.alignment = center_alignment # edit alignment
            
                data_subcategory_cell = row[2] # Data sub category row
                data_subcategory_cell.font = non_bold_font # edit font
                data_subcategory_cell.alignment = center_alignment # edit alignment

                inclusion_year_cell = row[3]
                inclusion_year_cell.font = non_bold_font # edit font
                inclusion_year_cell.alignment = center_alignment # edit alignment
 
                file_name_cell = row[4]
                file_name_cell.font = non_bold_font # edit font
                file_name_cell.alignment = right_alignment
            
            
        elif sheetname == "Variable Listings":
            listings = file_specs["Variable Listings"]
            
            for row in listings[3:listings.max_row]: # skip the headers
                
                site_cell = row[0]    # site column
                site_cell.alignment = center_alignment
                site_cell.font = bold_font
            
                data_category_cell = row[1] # data category column
                data_category_cell.font = bold_font # edit font
                data_category_cell.alignment = center_alignment # edit alignment
            
                data_subcategory_cell = row[2] # Data sub category row
                data_subcategory_cell.font = non_bold_font # edit font
                data_subcategory_cell.alignment = center_alignment # edit alignment

                inclusion_year_cell = row[3] # inclusion year(s)
                inclusion_year_cell.font = non_bold_font # edit font
                inclusion_year_cell.alignment = center_alignment # edit alignment
 
                file_name_cell = row[4] # file name
                file_name_cell.font = non_bold_font # edit font
                file_name_cell.alignment = right_alignment # edit alignment
            
                name_cell = row[5] # name
                name_cell.font = non_bold_font # edit font
                name_cell.alignment = center_alignment # edit alignment

                type_cell = row[6] # type
                type_cell.font = non_bold_font # edit font
                type_cell.alignment = center_alignment # edit alignment
            
                identifier_indicator_cell = row[7] #  indentifier indictor
                identifier_indicator_cell.font = non_bold_font # edit font
                identifier_indicator_cell.alignment = center_alignment # edit alignment
            
        else:
            next # skip if it is not the right sheet.
            
    writer.save()
    print("Template Saved")
    
# Constants 
# Using this dictionary because review file site names might be different than the folders
sjc_sites_dict = {"Ada": "Ada", "Allegheny": "Allegheny" , "Buncombe": "Buncombe", 
             "Charleston": "Charleston", "Cook": "Cook", "Harris": "Harris", "Lucas": "Lucas",
             "Mecklenburg": "Mecklenburg", "Milwaukee": "Milwaukee", "Multnomah": "Multnomah", 
             "New Orleans":"NOLA", "Palm Beach County":"PBC",
             "Pennington":"Pennington", "Philadelphia":"Philadelphia", 
             "Pima":"Pima", "Spokane": "Spokane", "St. Louis": "StLouis", "San Francisco":"SF", 
             "East Baton Rouge": "East Baton Rouge", "Lake": "Lake", "Minnehaha": "Minnehaha", 
             "Missoula": "Missoula"}


sjc_sites = ["Ada", "Allegheny", "Buncombe", "Charleston", "Cook", "Harris", "Lucas",
             "Mecklenburg", "Milwaukee", "Multnomah", "New Orleans", "Palm Beach County",
             "Pennington", "Philadelphia", "Pima", "Spokane", "St. Louis", "San Francisco", 
             "East Baton Rouge", "Lake", "Minnehaha", "Missoula"]


# import repository config file to get review file paths
repo_config = pd.read_excel("R:\SJC\Site information\Data Processing\Data Repository\CONFIG\SJC REPOSITORY CONFIG.xlsx", 
                           sheet_name = "Repository Paths")


file_path = site_file_path() # get file path to external request


data_df = populate_data_df(file_path) # get system points for external request

# function to populate file_descriptions and variable listings dataframes
file_descriptions, variable_listings = get_descriptions_and_listings(data_df)

file_descriptions = clean_descriptions_and_listings(file_descriptions)
variable_listings = clean_descriptions_and_listings(variable_listings)

# Function to save file_descriptions and variable listings to template
file_path = populate_specs_workbook(file_descriptions, variable_listings, file_path)

format_workbook_columns(file_path)

